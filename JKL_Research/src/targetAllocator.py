'''
Created on Aug 1, 2016

@author: kevin
'''
import openpyxl, pprint, os, yahoo_finance, numbers, datetime
from math import floor

if __name__ == '__main__':
    #TODO: Import both target allocation and current allocation weights.
    univList = ['VTI','QQQ','VXUS','VWO','REET','VNQ','BND','CMF','JNK','Cash']
    fundList = ['Aggressive','Moderate','Conservative']
    print('Reading Current Allocations...')
    curfile = os.path.join(os.path.abspath('..'),'AllocationFiles','CurrentKL.xlsx')
    curWb = openpyxl.load_workbook(curfile)
    curSheet = curWb.get_sheet_by_name('Sheet1')
    curAllocData = {}
    totalMv = 0
    for rowNum in range (2, curSheet.max_row+1):
    #TODO: Each Row represent a ticker and its weight
        ticker = curSheet.cell(row=rowNum,column = 1).value
        marketValue = curSheet.cell(row = rowNum,column = 2).value
        if not isinstance(marketValue, numbers.Number):
            continue
        curAllocData.setdefault(ticker,0)
        curAllocData[ticker] = marketValue
        totalMv += marketValue
    
    print ('Total MV',totalMv)
    
    print('Opening Allocation workbook...')
    tgtfile = os.path.join(os.path.abspath('..'),'AllocationFiles','TargetKL.xlsx')
    tgtWb = openpyxl.load_workbook(tgtfile)
    sheet = tgtWb.get_sheet_by_name('Sheet1')
    allocationData = {}
    # TODO: Fill in allocationData for each risk portfolio.  
    for col in range(2,sheet.max_column+1):
        # Each col in the spreadsheet is a portfolio
        portfolio = sheet.cell(row=1,column=col).value
        print(portfolio)
        totalTargetWt = 0
        totalTgtMV = 0
        for rowNum in range (2, sheet.max_row+1):
        #TODO: Each Row represent a ticker and its weight
            ticker = sheet.cell(row=rowNum,column = 1).value
            weights = sheet.cell(row = rowNum,column = col).value
            if not isinstance(weights, numbers.Number):
                continue
            allocationData.setdefault(portfolio,{})
            allocationData[portfolio].setdefault(ticker,{'weights':0,'MV':0,'Price':0,'OrderMV':0,'Shares':0,'NewMV':0,'NewWt':0})
            allocationData[portfolio][ticker]['weights'] = weights
            allocationData[portfolio][ticker]['MV'] = totalMv*weights
            tickerShare = yahoo_finance.Share(ticker)
            price = round(float(tickerShare.get_price()),2)
            allocationData[portfolio][ticker]['Price'] = price
            if ticker in curAllocData:
                currMV = curAllocData[ticker]
            else:
                currMV = 0
            orderMV = allocationData[portfolio][ticker]['MV']-currMV
            allocationData[portfolio][ticker]['OrderMV'] = orderMV
            allocationData[portfolio][ticker]['Shares'] = floor(float(orderMV)/price)
            allocationData[portfolio][ticker]['NewMV'] = round(currMV +float(allocationData[portfolio][ticker]['Shares'])*price,2) 
            allocationData[portfolio][ticker]['NewWt'] = round(allocationData[portfolio][ticker]['NewMV']/totalMv,4) 
            totalTgtMV += allocationData[portfolio][ticker]['NewMV']
            totalTargetWt += weights
        cashMV = round(totalMv - totalTgtMV,2)
        if cashMV >0 : 
            allocationData[portfolio].setdefault('Cash',{'weights':0,'MV':0,'Price':0,'OrderMV':0,'Shares':0,'NewMV':0})
            allocationData[portfolio]['Cash']['weights'] = 0
            allocationData[portfolio]['Cash']['MV'] = 0
            allocationData[portfolio]['Cash']['Price'] = 1
            allocationData[portfolio]['Cash']['Share'] = cashMV
            allocationData[portfolio]['Cash']['OrderMV'] = 0
            allocationData[portfolio]['Cash']['NewMV'] = cashMV
            allocationData[portfolio]['Cash']['NewWt'] = round(cashMV/totalMv,4)
            
        else:
            print('cash value is negative : ',cashMV)
            
        print (portfolio, ' Total Weights: ',totalTargetWt)
        if abs(totalTargetWt-1)>0.00001:
            print (portfolio,' Total Weights doesn''t equal to 1.')
    
    #TODO: Convert Weights to Shares
        
    
    #TODO: Convert Weights to Shares
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(allocationData)
    #TODO: Write out orders in a new sheet. 
    colHeader = ['weights','MV','Price','OrderMV','Shares','NewMV','NewWt'];
    for funds in fundList:
        tgtWb.create_sheet(title=funds, index=fundList.index(funds))
        tgtSheet = tgtWb.get_sheet_by_name(funds)
        # Documnt summary
        tgtSheet.cell(row=1,column=1).value='Total Asset'
        tgtSheet.cell(row=1,column=2).value=totalMv
        tgtSheet.cell(row=2,column=1).value='Date'
        tgtSheet.cell(row=2,column=2).value =datetime.date.today()
        tgtSheet.cell(row=3,column=1).value='Ticker'
        rowOffset = 3
        colOffset = 2
        for header in colHeader:
            tgtSheet.cell(row=rowOffset,column=colHeader.index(header)+colOffset).value = header
        for curTicker in univList:
            tgtSheet.cell(row=univList.index(curTicker)+rowOffset+1,column=1).value = curTicker
            for header in colHeader:
                tgtSheet.cell(row=univList.index(curTicker)+rowOffset+1,column = colHeader.index(header)+colOffset).value = allocationData[funds][curTicker][header]
     
    resultFile = os.path.join(os.path.abspath('..'),'AllocationFiles','result.xlsx')
    tgtWb.save(resultFile)
    print('Done')